from django.shortcuts import render, redirect
from datetime import date, timedelta
from calendar import monthrange
from django.db.models import Q
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font
import pandas as pd
import json

from helpers.utils import (
    parse_month,
    get_employees_queryset,
    get_schedule_maps,
    build_service_data,
    get_working_days,
)
from services.payroll import (
    count_shifts,
    calculate_shift_salary,
    calculate_service_sum,
)
from utils.excel_export import (
    workbook_to_response,
    autofit_columns,
    build_timesheet_workbook,
)

from .models import (
    Employee,
    WorkSchedule,
    ShiftType,
    Service,
    EmployeeServiceRecord,
    Department,
    ScheduleTemplate,
)


def timesheet_view(request):

    first_day = parse_month(request)
    year, month = first_day.year, first_day.month
    days_in_month = monthrange(year, month)[1]
    today = date.today()

    day_names = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    days_info = []
    for d in range(1, days_in_month + 1):
        dt = date(year, month, d)
        days_info.append({
            'num': d,
            'weekday': day_names[dt.weekday()],
            'is_today': dt == today,
            'is_weekend': dt.weekday() >= 5,
        })

    department_id = request.GET.get("department")
    employees = get_employees_queryset().order_by("department__name", "full_name")
    if department_id:
        employees = employees.filter(department_id=department_id)

    if request.method == "POST":
        for employee in employees:
            for day in range(1, days_in_month + 1):
                field_name = f"shift_{employee.id}_{day}"
                shift_value = request.POST.get(field_name)
                if shift_value:
                    date_obj = date(year, month, day)
                    WorkSchedule.objects.update_or_create(
                        employee=employee,
                        date=date_obj,
                        defaults={"shift": shift_value}
                    )
        redirect_url = f"{request.path}?month={first_day.strftime('%Y-%m')}"
        if department_id:
            redirect_url += f"&department={department_id}"
        return redirect(redirect_url)

    schedule_map, last_shift_map = get_schedule_maps(first_day)

    salary_summary = {}
    for employee in employees:
        shifts = schedule_map.get(employee.id, {})
        day_count, night_count = count_shifts(shifts)
        total_salary = calculate_shift_salary(employee, day_count, night_count)

        salary_summary[employee.id] = {
            "day": day_count,
            "night": night_count,
            "total": float(total_salary),
        }

    departments = Department.objects.all()
    templates = list(ScheduleTemplate.objects.all())
    templates_json = {t.id: t.sequence for t in templates}
    return render(request, "core/timesheet.html", {
        "employees": employees,
        "days": range(1, days_in_month + 1),
        "days_info": days_info,
        "month": first_day,
        "schedule": schedule_map,
        "salary": salary_summary,
        "shift_choices": [
        ("day", "Д"),
        ("night", "Н"),
        ("weekend", "В"),
        ("vacation", "О"),
        ("sick", "Б"),
        ("partial", "П"),
        ],
        "departments": departments,
        "selected_department": department_id,
        "last_shifts": last_shift_map,
        "schedule_templates": templates,
        "templates_json": templates_json,
    })


@require_POST
def apply_schedule_bulk(request):
    """Apply the same shift to all employees in a department for a date range."""
    try:
        data = json.loads(request.body.decode())
        month_str = data.get("month")
        department_id = data.get("department_id")
        shift = data.get("shift")
        start_day = int(data.get("start_day"))
        end_day = int(data.get("end_day"))

        if not all([month_str, department_id, shift]):
            return JsonResponse({"status": "error", "message": "Invalid data"}, status=400)

        year, month = [int(p) for p in month_str.split("-")]
        employees = Employee.objects.filter(department_id=department_id)
        for emp in employees:
            for d in range(start_day, end_day + 1):
                WorkSchedule.objects.update_or_create(
                    employee=emp,
                    date=date(year, month, d),
                    defaults={"shift": shift},
                )
        return JsonResponse({"status": "ok"})
    except Exception as exc:
        return JsonResponse({"status": "error", "message": str(exc)}, status=400)


def export_timesheet_xlsx(request):
    first_day = parse_month(request)
    wb = build_timesheet_workbook(first_day)
    filename = f"tabel_{first_day.year}_{first_day.month:02d}.xlsx"
    return workbook_to_response(wb, filename)


def send_timesheet_email(request):
    """Send timesheet Excel report to configured recipients."""
    first_day = parse_month(request)
    wb = build_timesheet_workbook(first_day)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    subject = f"Табель за {first_day.strftime('%B %Y')}"
    email = EmailMessage(
        subject=subject,
        body="Во вложении табель." ,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=settings.REPORT_RECIPIENTS,
    )
    filename = f"tabel_{first_day.year}_{first_day.month:02d}.xlsx"
    email.attach(filename, buffer.getvalue(),
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    email.send(fail_silently=False)

    messages.success(request, "Табель отправлен по электронной почте")
    return redirect(f"{reverse('report')}?month={first_day.strftime('%Y-%m')}")


def import_timesheet_view(request):
    """Import work schedules from uploaded Excel file."""
    first_day = parse_month(request)
    year, month = first_day.year, first_day.month
    days_in_month = monthrange(year, month)[1]

    if request.method == "POST" and request.FILES.get("xlsx_file"):
        wb = openpyxl.load_workbook(request.FILES["xlsx_file"])
        ws = wb.active

        shift_map = {
            "Д": "day",
            "д": "day",
            "Н": "night",
            "н": "night",
            "В": "weekend",
            "в": "weekend",
            "О": "vacation",
            "о": "vacation",
            "Б": "sick",
            "б": "sick",
            "П": "partial",
            "п": "partial",
            "Нп": "partial",
            "нп": "partial",
            "НП": "partial",
        }

        for row in ws.iter_rows(min_row=2):
            full_name = str(row[0].value).strip() if row[0].value else None
            if not full_name:
                continue
            try:
                employee = Employee.objects.get(full_name=full_name)
            except Employee.DoesNotExist:
                continue

            for day in range(1, min(days_in_month, len(row) - 1) + 1):
                raw = row[day].value
                if raw is None:
                    continue
                shift = shift_map.get(str(raw).strip())
                if not shift:
                    continue
                date_obj = date(year, month, day)
                WorkSchedule.objects.update_or_create(
                    employee=employee,
                    date=date_obj,
                    defaults={"shift": shift},
                )

        return redirect(f"{request.path}?month={first_day.strftime('%Y-%m')}")

    return render(request, "core/import_timesheet.html", {"month": first_day})

def services_view(request):
    first_day = parse_month(request)
    year, month = first_day.year, first_day.month
    employees = get_employees_queryset()
    services = Service.objects.all()
    selected_department = request.GET.get("department")

    if selected_department:
        employees = employees.filter(department_id=selected_department)

    if request.method == "POST":
        for employee in employees:
            # Используем фильтр по типу сотрудника
            filtered_services = services.filter(for_salary_based=employee.is_fixed_salary)
            for service in filtered_services:
                field = f"s_{employee.id}_{service.id}"
                qty = request.POST.get(field)
                if qty:
                    qty = int(qty)
                    EmployeeServiceRecord.objects.update_or_create(
                        employee=employee,
                        service=service,
                        month=first_day,
                        defaults={"quantity": qty}
                    )
        return redirect(f"{request.path}?month={first_day.strftime('%Y-%m')}&department={selected_department or ''}")

    records = EmployeeServiceRecord.objects.filter(month=first_day)
    data = build_service_data(records)

    summary = {}
    for employee in employees:
        filtered_services = services.filter(for_salary_based=employee.is_fixed_salary)
        quantities = {sid: data.get((employee.id, sid), 0) for sid in [s.id for s in filtered_services]}
        summary[employee.id] = float(calculate_service_sum(filtered_services, quantities))

    departments = Department.objects.all()

    return render(request, "core/services.html", {
        "employees": employees,
        "services": services,
        "month": first_day,
        "data": data,
        "summary": summary,
        "departments": departments,
        "selected_department": selected_department,
    })
def export_services_xlsx(request):
    first_day = parse_month(request)
    year, month = first_day.year, first_day.month

    employees = get_employees_queryset()
    services = Service.objects.all()
    records = EmployeeServiceRecord.objects.filter(month=first_day)

    raw_data = build_service_data(records)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Услуги_{year}_{month:02d}"

    headers = ["ФИО", "Отдел", "Должность"] + [s.name for s in services] + ["Итого (₽)"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for emp in employees:
        row = [emp.full_name, emp.department.name, emp.position.name]
        quantities = {s.id: raw_data.get((emp.id, s.id), 0) for s in services}
        for s in services:
            row.append(quantities.get(s.id, 0))
        total = calculate_service_sum(services, quantities)
        row.append(round(float(total), 2))
        ws.append(row)

    autofit_columns(ws)
    filename = f"uslugi_{year}_{month:02d}.xlsx"
    return workbook_to_response(wb, filename)


def import_services_view(request):
    """Import employee service records from an Excel file."""
    if request.method == "POST" and request.FILES.get("xlsx_file"):
        wb = openpyxl.load_workbook(request.FILES["xlsx_file"])
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        service_names = headers[2:]
        service_map = {
            s.name: s for s in Service.objects.filter(name__in=service_names)
        }
        for row in ws.iter_rows(min_row=2, values_only=True):
            full_name = row[0]
            month_val = row[1]
            if not full_name or not month_val:
                continue
            try:
                year, month = map(int, str(month_val).split("-")[:2])
                month_date = date(year, month, 1)
            except Exception:
                continue
            try:
                employee = Employee.objects.get(full_name=full_name)
            except Employee.DoesNotExist:
                continue
            for idx, service_name in enumerate(service_names, start=2):
                qty = row[idx]
                if qty:
                    service = service_map.get(service_name)
                    if not service:
                        continue
                    EmployeeServiceRecord.objects.update_or_create(
                        employee=employee,
                        service=service,
                        month=month_date,
                        defaults={"quantity": int(qty)},
                    )
        return redirect("services")
    return render(request, "core/import_services.html")


def export_salary_report_xlsx(request):
    first_day = parse_month(request)
    year, month = first_day.year, first_day.month
    days_in_month = monthrange(year, month)[1]

    employees = get_employees_queryset()
    records = EmployeeServiceRecord.objects.filter(month=first_day)
    services = Service.objects.all()

    schedule_map, _ = get_schedule_maps(first_day)
    service_data = build_service_data(records)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Зарплата_{year}_{month:02d}"

    headers = [
        "ФИО", "Отдел", "Должность",
        "Дневных", "Ночных", "Сумма смен (₽)", "Сумма услуг (₽)", "Итого (₽)"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    current_dept = None
    dept_day = dept_night = dept_shift_sum = dept_service_sum = dept_total = 0
    for emp in employees:
        if current_dept and emp.department_id != current_dept.id:
            ws.append([
                "",
                f"Итого по отделу {current_dept.name}",
                "",
                dept_day,
                dept_night,
                round(dept_shift_sum, 2),
                round(dept_service_sum, 2),
                round(dept_total, 2),
            ])
            dept_day = dept_night = dept_shift_sum = dept_service_sum = dept_total = 0

        shifts = schedule_map.get(emp.id, {})
        day_count, night_count = count_shifts(shifts)
        salary_shifts = calculate_shift_salary(emp, day_count, night_count)

        quantities = service_data.get(emp.id, {})
        service_sum = calculate_service_sum(services, quantities)

        total = round(float(salary_shifts + service_sum), 2)

        ws.append([
            emp.full_name,
            emp.department.name,
            emp.position.name,
            day_count,
            night_count,
            round(float(salary_shifts), 2),
            round(float(service_sum), 2),
            total,
        ])

        dept_day += day_count
        dept_night += night_count
        dept_shift_sum += float(salary_shifts)
        dept_service_sum += float(service_sum)
        dept_total += total
        current_dept = emp.department

    if current_dept:
        ws.append([
            "",
            f"Итого по отделу {current_dept.name}",
            "",
            dept_day,
            dept_night,
            round(dept_shift_sum, 2),
            round(dept_service_sum, 2),
            round(dept_total, 2),
        ])

    autofit_columns(ws)
    filename = f"zarplata_{year}_{month:02d}.xlsx"
    return workbook_to_response(wb, filename)


def report_view(request):
    first_day = parse_month(request)
    selected_department = request.GET.get("department")
    report_type = request.GET.get("type")  # "advance" или "final"

    year, month = first_day.year, first_day.month
    days_in_month = monthrange(year, month)[1]
    employees = get_employees_queryset().order_by("department__name", "full_name")
    if selected_department:
        employees = employees.filter(department_id=selected_department)

    schedule_map, last_shift_map = get_schedule_maps(first_day)

    services = Service.objects.all()
    records = EmployeeServiceRecord.objects.filter(month=first_day)
    service_data = build_service_data(records)

    salary_summary = []
    department_totals: dict[int, dict[str, float]] = {}
    working_days_total = get_working_days(year, month)
    mid_month = 15

    for emp in employees:
        shifts = schedule_map.get(emp.id, {})
        day_count, night_count = count_shifts(shifts)
        total_shift_salary = calculate_shift_salary(emp, day_count, night_count)

        quantities = service_data.get(emp.id, {})
        service_sum = calculate_service_sum(services, quantities)

        if emp.is_fixed_salary:
            worked_days = sum(1 for d, s in shifts.items() if s in ["day", "night"])
            salary_base = (float(emp.fixed_salary) / working_days_total) * worked_days if working_days_total else 0
            total_salary = salary_base + float(emp.bonus) + float(service_sum)

            if report_type == "advance":
                worked_days_1_15 = sum(
                    1 for d, s in shifts.items() if s in ["day", "night"] and d <= mid_month
                )
                advance = (
                    (float(emp.fixed_salary) / working_days_total) * worked_days_1_15
                    if working_days_total
                    else 0
                )
                row = {
                    "employee": emp,
                    "day": worked_days_1_15,
                    "night": 0,
                    "shifts_sum": round(advance),
                    "services_sum": 0,
                    "total": round(advance),
                }
                salary_summary.append(row)
                totals = department_totals.setdefault(
                    emp.department_id,
                    {"day": 0, "night": 0, "total": 0},
                )
                totals["day"] += row["day"]
                totals["night"] += row["night"]
                totals["total"] += row["total"]
                continue
        else:
            total_salary = float(total_shift_salary + service_sum)

        row = {
            "employee": emp,
            "day": day_count,
            "night": night_count,
            "shifts_sum": round(float(total_shift_salary)),
            "services_sum": round(float(service_sum)),
            "total": round(total_salary),
        }
        salary_summary.append(row)
        totals = department_totals.setdefault(
            emp.department_id,
            {"day": 0, "night": 0, "total": 0},
        )
        totals["day"] += row["day"]
        totals["night"] += row["night"]
        totals["total"] += row["total"]

    return render(request, "core/report.html", {
        "month": first_day,
        "departments": Department.objects.all(),
        "selected_department": selected_department,
        "report_type": report_type,
        "salary_summary": salary_summary,
        "department_totals": department_totals,
    })

def export_salary_full_xlsx(request):
    employees = get_employees_queryset()
    return generate_salary_report(request, employees, full_month=True)

def export_salary_advance_xlsx(request):
    employees = get_employees_queryset()
    return generate_salary_report(request, employees, full_month=False)

def generate_salary_report(request, employees, full_month=True):
    from calendar import monthrange
    from openpyxl import Workbook
    from openpyxl.styles import Font

    first_day = parse_month(request)
    year, month = first_day.year, first_day.month
    days_in_month = monthrange(year, month)[1]

    selected_department = request.GET.get("department")
    if selected_department:
        employees = employees.filter(department_id=selected_department)

    schedule_map, _ = get_schedule_maps(first_day)
    records = EmployeeServiceRecord.objects.filter(month=first_day)
    services = Service.objects.all()

    service_data = build_service_data(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "Аванс" if not full_month else "Зарплата"

    headers = ["ФИО", "Отдел", "Должность", "Дневных", "Ночных", "Сумма смен (₽)", "Сумма услуг (₽)", "Премия", "Итого (₽)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for emp in employees:
        shifts = schedule_map.get(emp.id, {})

        if full_month:
            period_days = range(1, days_in_month + 1)
        else:
            period_days = range(1, 16)  # с 1 по 15

        day_count, night_count = count_shifts(shifts, period_days)

        if emp.is_fixed_salary:
            total_days = sum(1 for d in range(1, days_in_month + 1) if shifts.get(d) in ["day", "night"])
            worked_days = sum(1 for d in period_days if shifts.get(d) in ["day", "night"])
            base = (emp.fixed_salary or 0) / total_days if total_days else 0
            salary_shift_sum = round(base * worked_days, 2)
        else:
            salary_shift_sum = round(float(calculate_shift_salary(emp, day_count, night_count)), 2)

        quantities = service_data.get(emp.id, {})
        service_sum = calculate_service_sum(services, quantities)

        bonus = float(emp.bonus or 0) if full_month else 0
        total = Decimal(salary_shift_sum) + (Decimal(service_sum) if full_month else Decimal(0)) + Decimal(bonus)

        ws.append([
            emp.full_name,
            emp.department.name,
            emp.position.name,
            day_count,
            night_count,
            salary_shift_sum,
            round(service_sum if full_month else 0, 2),
            round(bonus, 2),
            round(total, 2)
        ])

    # автоширина
    autofit_columns(ws)

    filename = f"{'avans' if not full_month else 'zarplata'}_{year}_{month:02d}.xlsx"
    return workbook_to_response(wb, filename)


def analytics_view(request):
    """Сводная аналитика по сменам и услугам по отделам."""
    first_day = parse_month(request)
    year, month = first_day.year, first_day.month

    # Данные по сменам
    qs_shifts = (
        WorkSchedule.objects
        .filter(date__year=year, date__month=month)
        .select_related("employee__department")
    )
    df_shifts = pd.DataFrame.from_records(
        qs_shifts.values("employee__department__name", "shift")
    )
    if not df_shifts.empty:
        pivot_shifts = (
            df_shifts
            .groupby(["employee__department__name", "shift"])
            .size()
            .unstack(fill_value=0)
            .reset_index()
        )
        shift_chart = (
            pivot_shifts.set_index("employee__department__name").sum(axis=1).to_dict()
        )
        shift_table_html = pivot_shifts.to_html(
            index=False,
            classes="w-full border-collapse text-sm text-gray-800",
            border=0,
        )
    else:
        shift_chart = {}
        shift_table_html = ""

    # Данные по услугам
    qs_services = (
        EmployeeServiceRecord.objects
        .filter(month=first_day)
        .select_related("employee__department", "service")
    )
    df_services = pd.DataFrame.from_records(
        qs_services.values("employee__department__name", "service__name", "quantity")
    )
    if not df_services.empty:
        pivot_services = (
            df_services
            .groupby(["employee__department__name", "service__name"])["quantity"]
            .sum()
            .unstack(fill_value=0)
            .reset_index()
        )
        service_chart = (
            pivot_services.set_index("employee__department__name").sum(axis=1).to_dict()
        )
        service_table_html = pivot_services.to_html(
            index=False,
            classes="w-full border-collapse text-sm text-gray-800",
            border=0,
        )
    else:
        service_chart = {}
        service_table_html = ""

    context = {
        "month": first_day,
        "shift_table": shift_table_html,
        "service_table": service_table_html,
        "shift_chart": json.dumps(shift_chart, ensure_ascii=False),
        "service_chart": json.dumps(service_chart, ensure_ascii=False),
    }
    return render(request, "core/analytics.html", context)

