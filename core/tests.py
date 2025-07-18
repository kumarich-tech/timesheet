from datetime import date
from io import BytesIO
import json

from django.test import TestCase, RequestFactory
from django.urls import reverse
import openpyxl

from .models import (
    Department,
    Position,
    Employee,
    WorkSchedule,
    Service,
    EmployeeServiceRecord,
    ScheduleTemplate,
)

from helpers.utils import parse_month
from services.payroll import calculate_shift_salary, calculate_service_sum


class TimesheetViewTests(TestCase):
    def setUp(self):
        self.department = Department.objects.create(name="Dep")
        self.position = Position.objects.create(name="Worker")
        self.employee = Employee.objects.create(
            full_name="John Doe",
            department=self.department,
            position=self.position,
        )

    def test_get_timesheet(self):
        url = reverse("timesheet") + "?month=2024-01"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertIn(self.employee, list(response.context["employees"]))

    def test_post_timesheet_creates_schedule(self):
        month = date.today().strftime("%Y-%m")
        url = reverse("timesheet") + f"?month={month}"
        resp = self.client.post(url, {f"shift_{self.employee.id}_1": "day"})
        self.assertEqual(resp.status_code, 302)
        month_first = date.today().replace(day=1)
        self.assertTrue(
            WorkSchedule.objects.filter(
                employee=self.employee, date=month_first, shift="day"
            ).exists()
        )

    def test_schedule_templates_in_context(self):
        ScheduleTemplate.objects.create(name="2/2", sequence=["day", "day", "weekend", "weekend"])
        url = reverse("timesheet") + "?month=2024-01"
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("schedule_templates", resp.context)

    def test_apply_schedule_bulk(self):
        emp2 = Employee.objects.create(full_name="Jane", department=self.department, position=self.position)
        month = date.today().strftime("%Y-%m")
        url = reverse("apply_schedule_bulk")
        payload = {
            "department_id": self.department.id,
            "shift": "day",
            "start_day": 1,
            "end_day": 2,
            "month": month,
        }
        resp = self.client.post(url, data=json.dumps(payload), content_type="application/json")
        self.assertEqual(resp.status_code, 200)
        first_day = date.today().replace(day=1)
        for emp in [self.employee, emp2]:
            for d in range(1, 3):
                self.assertTrue(
                    WorkSchedule.objects.filter(employee=emp, date=date(first_day.year, first_day.month, d), shift="day").exists()
                )


class ServicesViewTests(TestCase):
    def setUp(self):
        self.department = Department.objects.create(name="Dep")
        self.position = Position.objects.create(name="Worker")
        self.employee = Employee.objects.create(
            full_name="John Doe",
            department=self.department,
            position=self.position,
            is_fixed_salary=False,
        )
        self.service = Service.objects.create(
            name="Test Service", price=100, for_salary_based=False
        )

    def test_get_services(self):
        url = reverse("services") + "?month=2024-01"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertIn(self.service, list(response.context["services"]))

    def test_post_services_creates_record(self):
        month = date.today().strftime("%Y-%m")
        url = reverse("services") + f"?month={month}"
        resp = self.client.post(
            url, {f"s_{self.employee.id}_{self.service.id}": "2"}
        )
        self.assertEqual(resp.status_code, 302)
        month_first = date.today().replace(day=1)
        record = EmployeeServiceRecord.objects.get(
            employee=self.employee, service=self.service, month=month_first
        )
        self.assertEqual(record.quantity, 2)


class ImportServicesTests(TestCase):
    def setUp(self):
        self.department = Department.objects.create(name="Dep")
        self.position = Position.objects.create(name="Worker")
        self.employee = Employee.objects.create(
            full_name="John Doe",
            department=self.department,
            position=self.position,
        )
        self.service_a = Service.objects.create(name="Service A", price=10)
        self.service_b = Service.objects.create(name="Service B", price=20)

    def _create_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ФИО", "Месяц", self.service_a.name, self.service_b.name])
        ws.append(["John Doe", "2024-01", 1, 2])
        return wb

    def test_import_services_creates_records(self):
        wb = self._create_workbook()
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        resp = self.client.post(reverse("import_services"), {"xlsx_file": buffer}, format="multipart")
        self.assertEqual(resp.status_code, 302)
        month = date(2024, 1, 1)
        rec_a = EmployeeServiceRecord.objects.get(employee=self.employee, service=self.service_a, month=month)
        rec_b = EmployeeServiceRecord.objects.get(employee=self.employee, service=self.service_b, month=month)
        self.assertEqual(rec_a.quantity, 1)
        self.assertEqual(rec_b.quantity, 2)


class ParseMonthTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_valid_month(self):
        request = self.factory.get("/", {"month": "2024-05"})
        result = parse_month(request)
        self.assertEqual(result, date(2024, 5, 1))

    def test_invalid_month_returns_current(self):
        today = date.today()
        request = self.factory.get("/", {"month": "invalid"})
        result = parse_month(request)
        self.assertEqual(result, date(today.year, today.month, 1))

    def test_no_month_returns_current(self):
        today = date.today()
        request = self.factory.get("/")
        result = parse_month(request)
        self.assertEqual(result, date(today.year, today.month, 1))


class ExcelExportTests(TestCase):
    def setUp(self):
        self.department = Department.objects.create(name="Dep")
        self.position = Position.objects.create(name="Worker")
        self.emp1 = Employee.objects.create(
            full_name="John Doe", department=self.department, position=self.position
        )
        self.emp2 = Employee.objects.create(
            full_name="Jane Smith", department=self.department, position=self.position
        )
        self.service = Service.objects.create(name="Test Service", price=100)

    def _load_workbook(self, response):
        return openpyxl.load_workbook(BytesIO(response.content))

    def test_timesheet_export_dimensions(self):
        url = reverse("export_timesheet") + "?month=2024-01"
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        wb = self._load_workbook(resp)
        ws = wb.active
        self.assertEqual(ws.max_row, Employee.objects.count() + 1)
        self.assertEqual(
            ws.max_column,
            3 + 31 + 3,  # 2024-01 has 31 days
        )

    def test_services_export_dimensions(self):
        url = reverse("export_services") + "?month=2024-01"
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        wb = self._load_workbook(resp)
        ws = wb.active
        self.assertEqual(ws.max_row, Employee.objects.count() + 1)
        self.assertEqual(ws.max_column, 3 + Service.objects.count() + 1)

    def test_salary_report_export_dimensions(self):
        url = reverse("export_salary_report") + "?month=2024-01"
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        wb = self._load_workbook(resp)
        ws = wb.active
        dept_count = Department.objects.count()
        self.assertEqual(ws.max_row, Employee.objects.count() + dept_count + 1)
        self.assertEqual(ws.max_column, 8)


class PayrollCalculationTests(TestCase):
    def setUp(self):
        self.department = Department.objects.create(name="Dep")
        self.position = Position.objects.create(name="Worker")
        self.fixed_emp = Employee.objects.create(
            full_name="Fixed", department=self.department, position=self.position,
            is_fixed_salary=True, fixed_salary=30000, bonus=5000
        )
        self.piece_emp = Employee.objects.create(
            full_name="Piece", department=self.department, position=self.position,
            day_shift_rate=1000, night_shift_rate=1500
        )
        self.service_a = Service.objects.create(name="A", price=100)
        self.service_b = Service.objects.create(name="B", price=50)

    def test_fixed_employee_salary(self):
        salary = calculate_shift_salary(self.fixed_emp, 5, 3)
        self.assertEqual(salary, 35000)

    def test_piece_employee_salary(self):
        salary = calculate_shift_salary(self.piece_emp, 2, 1)
        self.assertEqual(salary, 3500)

    def test_calculate_service_sum(self):
        qty = {self.service_a.id: 3, self.service_b.id: 2}
        result = calculate_service_sum(Service.objects.all(), qty)
        self.assertEqual(result, 400)

